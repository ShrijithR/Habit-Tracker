#!Habit track Functions

import os
import logging
import importlib
import openpyxl as xl
import datetime as dt
#                                           importing datetime modules of both python 3.6 and 3.8
#                                           to use 'fromisocalendar' attribute of python 3.8 unavailable in 3.6.
#                                           currently working IDE handles only the version 3.6
import datetime38 as dt38
import pyinputplus as pyip
import pprint as pp
import HabitGoalsFile as hgf
from HabitListFile import habitList

logging.basicConfig(level=logging.INFO,format='%(message)s')
logging.disable(logging.INFO)
logging.disable(logging.ERROR)
logging.disable(logging.CRITICAL)
"""Check the current working directory for files that start with the
name 'habit tracker' and ends with the format '.xlsx'
If multiple files are present, a file is chosen from the list or else
the single file is chosen or returns no files present
"""
workbookfiles = []
for xlFile in os.listdir(os.getcwd()):
    if xlFile.startswith("Habit Tracker") and xlFile.endswith(".xlsx"):
        workbookfiles.append(xlFile)
try:
    workbookChoice = workbookfiles[0]
except IndexError:
    print("no files found")
if len(workbookfiles) > 1:
    print("Choose the tracker workbook to be updated")
    workbookChoice = pyip.inputMenu(workbookfiles, numbered=True)


#                                           LOADING WORKBOOK AND SHEET OBJECT
wb = xl.load_workbook(workbookChoice)
sheetHabit = wb["HABITS"]
#                                           FUTURE IMPLEMENTATION - Add sleep and expenditure tracker sheets
#                                           options for choosing from the available sheets

#                                           START DATE AND END DATE
#                                           Calculate the date from which the sheet begins and ends
#                                           in order to check whether the input date falls between

#                                           Add week names from the sheet to a list
#                                           to find and choose the first and the last week
weekList = []
for i in range(1, sheetHabit.max_column + 1):
    weekValue = sheetHabit.cell(row=2, column=i).value
    if str(weekValue).lower().startswith("w"):
        weekList.append(weekValue)
weekStartNum = int(str(weekList[0][1:]))
weekEndNum = int(str(weekList[(len(weekList) - 1)][1:]))
#                                           Find the starting day and the ending day of the sheet
#                                           Compare it to a list containing the names of days
#                                           to assign the index of the elements to the day names
#                                           thereby getting the day number
dayNameList = ["mon", "tue", "wed", "thur", "fri", "sat", "sun"]
dayStartName = sheetHabit.cell(row=3, column=3).value
dayStartNum = dayNameList.index(dayStartName) + 1
dayEndName = sheetHabit.cell(row=3, column=sheetHabit.max_column).value
dayEndNum = dayNameList.index(dayEndName) + 1
#                                           get the date from the isocalendar attribute which accepts parameters (year, week, daynumber)
#                                           Convert the date object into datetime.date object
#                                           to compare it with the input date
date38StartLimit = dt38.date.fromisocalendar(2020, weekStartNum, dayStartNum)
dateStartLimit = dt.date.fromordinal(date38StartLimit.toordinal())
date38EndLimit = dt38.date.fromisocalendar(2020, weekEndNum, dayEndNum)
dateEndLimit = dt.date.fromordinal(date38EndLimit.toordinal())
#                                           UPDATE HABIT DATABASE
#                                           Parse the rows of the second column
#                                           to write the habits in the list-'HabitList'
#                                           and create/edit the module-'HabitListFile'
def updateHabitList():
    listData = []
    for rws in range(1, sheetHabit.max_row + 1):
        habit = sheetHabit.cell(row=rws, column=2).value
        if habit:
            listData.append(habit)
    habitListFile = open("HabitListFile.py", "w")
    habitListFile.write("habitList=" + pp.pformat(listData))
    habitListFile.close()


'''Update the goals dictionary-'habitGoalsDict' for a single/all habits
Create/edit the same in habitGoalsFile module
FUTURE IMPLEMENTATION - modify multiple habits
'''
def updateGoals(goalForHabit):
    choiceChange = "yes"
    while choiceChange == "yes":
#                                           Check if the dictionary is present in the module
        try:
            for habitName in hgf.habitGoalsDict.keys():
                print(
                    f"{habitName} : "
                    f"{list(hgf.habitGoalsDict[habitName].keys())[0]} : "
                    f"{list(hgf.habitGoalsDict[habitName].values())[0]}"
                )
        except AttributeError:
            pass
        if goalForHabit == "all":
            habitsGoalsDict = {}
            for habit in habitList:
                habitsGoalsDict[habit] = {
                    "Goal": pyip.inputInt(f"Enter your goal for {habit}\n",min=1)
                }
            habitGoalsFile = open("HabitGoalsFile.py", "w")
            habitGoalsFile.write("habitGoalsDict=" + pp.pformat(habitsGoalsDict))
            habitGoalsFile.close()
        else:
            hgf.habitGoalsDict[goalForHabit] = {
                "Goal": pyip.inputInt(f"Enter your goal for {goalForHabit}\n",min=1)
            }
            habitGoalsFile = open("HabitGoalsFile.py", "w")
            habitGoalsFile.write("habitGoalsDict=" + pp.pformat(hgf.habitGoalsDict))
            habitGoalsFile.close()
#                                           Reload the module in order to update the dictionary
        importlib.reload(hgf)
        for habitName in hgf.habitGoalsDict.keys():
            print(
                f"{habitName} : "
                f"{list(hgf.habitGoalsDict[habitName].keys())[0]} : "
                f"{list(hgf.habitGoalsDict[habitName].values())[0]}"
            )
        choiceChange = pyip.inputYesNo(
            prompt="Do you want to modify the current goals?\n"
        )

#                                           FIND HABIT ROW
#                                           Find the row number of the habit parameter and return as integer

def findHabitrow(habitName):
    for rw in range(1, sheetHabit.max_row + 1):
        if habitName == str(sheetHabit.cell(row=rw, column=2).value).lower():
            return rw


#                                           FIND THE DATES
#                                           Find the beginning/ending dates of previous/current month/week
def dateFinder():
    import calendar as cal
    from dateutil.relativedelta import relativedelta as rd

    currDate = dt.date.today()
    currMonth = dt.date.today().month
    currYear = dt.date.today().year
    currDateDelta = dt.timedelta(days=(currDate.isocalendar()[2] - 1))
#                                           Finding the start of the current week by
#                                           subracting the day number of the week from the day of the month
    dateFinder.currWeekStart = currDate - currDateDelta
    dateFinder.currWeekEnd = dateFinder.currWeekStart + dt.timedelta(days=6)
    dateFinder.currMonthStart = dt.date.today().replace(day=1)
    dateFinder.currMonthEnd = dt.date.today().replace(
        day=(cal.monthrange(currYear, currMonth)[1])
    )
    prevWeekDelta = dt.timedelta(days=7)
    dateFinder.prevWeekStart = dateFinder.currWeekStart - prevWeekDelta
    dateFinder.prevWeekEnd = dateFinder.prevWeekStart + dt.timedelta(days=6)
    dateFinder.prevMonthStart = dateFinder.currMonthStart - rd(months=1)
    dateFinder.prevMonthEnd = dateFinder.prevMonthStart + dt.timedelta(
        days=(
        cal.monthrange(currYear, (dateFinder.prevMonthStart.month))[1] - 1)
    )
dateFinder()

#                                           Check if the date parameter is above the starting date of the tracker
#                                           and below the endign date of the tracker

def dateCheck(dateObj):
    if not dateObj < dateStartLimit or dateObj > dateEndLimit:
        return True
    print(
        "Date must be between %s/%d/%d and %s/%d/%d"
        % (dateStartLimit.day, dateStartLimit.month, dateStartLimit.year,
           dateEndLimit.day, dateEndLimit.month, dateEndLimit.year)
    )
    return False
  
#                                           Prompt the user for the date, run teh dateCheck function
#                                           and return the date if the function is True
def getDate(msg):
    while True:
        dateOutput=pyip.inputDate(
            prompt=msg,formats=(["%d/%m/%y"])
        )
        if dateCheck(dateOutput):
            break
    return dateOutput
#                                           Find the column number of the date parameter by parsing the rows of week number and the day number

def findColumnRange(targetDate):
    #                                       Find the column of the week number of the target date
    #                                       then find the target day across the following seven columns
    #                                       since there would exists the same day number in several months
    targetWeekNum = "w" + str(targetDate.isocalendar()[1])
    targetDayNum = targetDate.day
    for col in range(1, sheetHabit.max_column):
        if str(
        sheetHabit.cell(row=2, column=col).value).lower() == targetWeekNum:
            colWeek = col
            break
    for colDay in range(colWeek, colWeek + 7):
        if sheetHabit.cell(row=4, column=colDay).value == targetDayNum:
            colNum = colDay
            break
    return colNum


"""Get the input habits and dates from the user and store it in respective variables
"""


def infoReq():
    print("Select the habit you want to check. ", end="")
    habitChoice = pyip.inputMenu(habitList + ["all"], numbered=True)
    logging.error(habitChoice)
    print("Enter the dates you want to check. ", end="")
    dateList = [
        "Previous week",
        "Current week",
        "Previous month",
        "Current month",
        "Custom date"]
    dateChoice = pyip.inputMenu(dateList, numbered=True)
    if dateChoice == "Previous week":
        dateAObj = dateFinder.prevWeekStart
        dateBObj = dateFinder.prevWeekEnd
    elif dateChoice == "Current week":
        dateAObj = dateFinder.currWeekStart
        dateBObj = dateFinder.currWeekEnd
    elif dateChoice == "Previous month":
        dateAObj = dateFinder.prevMonthStart
        dateBObj = dateFinder.prevMonthEnd
    elif dateChoice == "Current month":
        dateAObj = dateFinder.currMonthStart
        dateBObj = dateFinder.currMonthEnd
    elif dateChoice == "Custom date":
        dateChoiceB = pyip.inputMenu(
        ["Single date", "Date range"], numbered=True)
        if dateChoiceB == "Date range":
            dateAObj=getDate("Enter the start date in the format dd/mm/yy\n")
            dateBObj=getDate("Enter the end date in the format dd/mm/yy\n")
        else:
            dateAObj=getDate("Enter the date in the format dd/mm/yy\n")
            dateBObj = dateAObj

    return habitChoice, dateAObj, dateBObj

"""Returns a dictionary containing the results of all the habits across the date range
"""


def analysis(dateA, dateB):
    habitsResultsDict = {}
    colStartRange = findColumnRange(dateA)
    colEndRange = findColumnRange(dateB)
    for habit in habitList:
        rw = findHabitrow(habit)
        for col in range(colStartRange, colEndRange + 1):
            habitResult = str(
            sheetHabit.cell(row=rw, column=col).value).lower()
            yesCount = 0
            noCount = 0
            if habitResult == "y":
                yesCount += 1
            elif habitResult == "n":
                noCount += 1
        habitsResultsDict[habit] = {"yes": yesCount, "no": noCount}
    logging.critical(habitsResultsDict)
    return habitsResultsDict


"""Writes the results of habits across two date ranges onto a text file
"""


def compare(habit, dateChoice):
    if dateChoice == "Previous week vs Current week":
        dateStartA = dateFinder.prevWeekStart
        dateEndA = dateFinder.prevWeekEnd
        dateStartB = dateFinder.currWeekStart
        dateEndB = dateFinder.currWeekEnd
    elif dateChoice == "Previous month vs current month":
        dateStartA = dateFinder.prevMonthStart
        dateEndA = dateFinder.prevMonthEnd
        dateStartB = dateFinder.currMonthStart
        dateEndB = dateFinder.currMonthEnd
    elif dateChoice == "Custom date range":
        print("Single dates or Date ranges? ", end="")
        dateChoiceB = pyip.inputMenu(
        ["Single dates", "Date ranges"], numbered=True)
        if dateChoiceB == "Date ranges":
            dateStartA = getDate("Enter the range A start date in the format dd/mm/yy\n")
            dateEndA = getDate("Enter the range A end date in the format dd/mm/yy\n")
            dateStartB = getDate("Enter the range B start date in the format dd/mm/yy\n")
            dateEndB = getDate("Enter the range B end date in the format dd/mm/yy\n")
        else:
            dateStartA = getDate("Enter the date A in the format dd/mm/yy\n")
            dateStartB = getDate("Enter the date B in the format dd/mm/yy\n")
            dateEndA = dateStartA
            dateEndB = dateStartB
    resultsDictA = analysis(dateStartA, dateEndA)
    resultsDictB = analysis(dateStartB, dateEndB)
    if habit == "all":
        compare.outputFileName = (f"Comparison of all habits between"
        f"{dateStartA.day}.{dateStartA.month}-{dateEndA.day}.{dateEndA.month} &"
        f"{dateStartB.day}.{dateStartB.month}-{dateEndB.day}.{dateEndB.month}.txt"
        )
        outputFile = open(
            os.path.join(
            os.path.abspath("OutputFiles"), compare.outputFileName
            ), "w"
        )
        outputFile.write(
            f"{'HABIT'.ljust(20,'-')}{str(dateStartA.day)}/"
            f"{str(dateStartA.month)}-{str(dateEndA.day)}/"
            f"{str(dateEndA.month).ljust(15,'-')}{str(dateStartB.day)}/"
            f"{str(dateStartB.month)}-{str(dateEndB.day)}/{str(dateEndB.month)}\n"
        )
        for key in resultsDictA.keys():
            if key not in resultsDictB.keys():
                return
            outputFile.write(
                f"{key.ljust(20)}Yes: {str(resultsDictA[key]['yes'])} "
                f"No: {str(resultsDictA[key]['no']).ljust(10)}"
                f"Yes: {str(resultsDictB[key]['yes'])} "
                f"No: {str(resultsDictB[key]['no'])}\n"
            )
        return
    compare.outputFileName = (f"Comparison of {habit} between "
    f"{dateStartA.day}.{dateStartA.month}-{dateEndA.day}.{dateEndA.month} & "
    f"{dateStartB.day}.{dateStartB.month}-{dateEndB.day}.{dateEndB.month}.txt"
    )
    outputFile = open(
        os.path.join(
        os.path.abspath("OutputFiles"), compare.outputFileName
        ), "w"
    )
    outputFile.write(
        f"{'HABIT'.ljust(20,'-')}{str(dateStartA.day)}/{str(dateStartA.month)}-"
        f"{str(dateEndA.day)}/{str(dateEndA.month).ljust(15,'-')}"
        f"{str(dateStartB.day)}/{str(dateStartB.month)}-"
        f"{str(dateEndB.day)}/{str(dateEndB.month)}\n"
    )
    if habit not in resultsDictA.keys() and habit not in resultsDictB.keys():
        return
    outputFile.write(
        f"{habit.ljust(20)}Yes: {str(resultsDictA[habit]['yes'])} "
        f"No: {str(resultsDictA[habit]['no']).ljust(10)}"
        f"Yes: {str(resultsDictB[habit]['yes'])} "
        f"No: {str(resultsDictB[habit]['no'])}\n"
    )


"""Checks if the goals have or haven't been reached for the habit parameter across the date range adn writes the output
onto a text file
"""
def goalcheck(habit, dateRange):
    dateStart = dateRange[0]
    dateEnd = dateRange[1]
    remDays = 0
    if dateEnd > dt.date.today():
        remDays = (dateEnd - dt.date.today()).days
    for habitName in hgf.habitGoalsDict.keys():
        print(
            f"{habitName} : "
            f"{list(hgf.habitGoalsDict[habitName].keys())[0]} : "
            f"{list(hgf.habitGoalsDict[habitName].values())[0]}"
        )
    choiceChange = pyip.inputYesNo(
    prompt="Do you want to modify the current goals?\n")
    if choiceChange == "yes":
        updateGoals(habit)
    dateDict = analysis(dateStart, dateEnd)
    if habit == "all":
        goalcheck.outputFileName = (f"Goal check of all habits between "
        f"{dateStart.day}.{dateStart.month}-{dateEnd.day}.{dateEnd.month}.txt"
        )
        outputFile = open(
            os.path.join(
            os.path.abspath("OutputFiles"), goalcheck.outputFileName), "w"
        )
        outputFile.write(
            f"{'HABIT'.ljust(15,'-')}"
            f"{dateStart.day}/{dateStart.month}-"
            f"{dateEnd.day}/{dateEnd.month}\n"
        )
        for habitName in dateDict.keys():
            if not habitName in hgf.habitGoalsDict.keys():
                return
            outputFile.write(
                f"{habitName.ljust(15)}"
                f"total days: {str((dateEnd-dateStart).days+1).ljust(10)}"
                f"Remaining days: {str(remDays).ljust(10)}\n"
                f"Goals: {str(list(hgf.habitGoalsDict[habitName].values())[0]).ljust(10)}"
                f"Number of times done: {str(dateDict[habitName]['yes']).ljust(5)}\n"
            )
            if (
                dateDict[habitName]["yes"]
                >= list(hgf.habitGoalsDict[habitName].values())[0]
            ):
                outputFile.write("Goal reached\n\n")
                return
            outputFile.write(
                f"{(list(hgf.habitGoalsDict[habitName].values())[0])-dateDict[habitName]['yes']} "
                f"days to reach the goal\n\n"
            )
        return
    goalcheck.outputFileName = (f"goal check of {habit} between "
    f"{dateStart.day}.{dateStart.month}-{dateEnd.day}.{dateEnd.month}.txt"
    )
    outputFile = open(
        os.path.join(
        os.path.abspath("OutputFiles"), goalcheck.outputFileName), "w"
    )
    outputFile.write(
        f"{'HABIT'.ljust(15,'-')}"
        f"{dateStart.day}/{str(dateStart.month).ljust(10,'-')}"
        f"{dateEnd.day}/{dateEnd.month}\n"
    )
    if habit not in dateDict.keys() and habit not in hgf.habitGoalsDict.keys():
        return
    outputFile.write(
        f"{habit.ljust(15)}total days: "
        f"{str((dateEnd-dateStart).days+1).ljust(10)}"
        f"Remaining days: {str(remDays).ljust(10)}\n"
        f"Goals: {str(list(hgf.habitGoalsDict[habit].values())[0]).ljust(10)}"
        f"Number of times done: {str(dateDict[habit]['yes']).ljust(5)}\n"
    )
    if (
        dateDict[habit]["yes"]
        >= list(hgf.habitGoalsDict[habit].values())[0]
    ):
        outputFile.write("Goal reached\n\n")
        return
    outputFile.write(
        f"{(list(hgf.habitGoalsDict[habit].values())[0])-dateDict[habit]['yes']} "
        f"days to reach the goal\n\n"
    )


"""Update a single/all habits
Create backup of the current tracker
Saves the updated tracker in its place
Delete excess backup files
"""
#                                           TO DO FUTURE - Be able to update select a specific number of habits

def update():
    import shutil

    today = dt.date.today()
    col = findColumnRange(today)
    print("Which habit do you want to update?", end=" ")
    habitToBeUpdated = pyip.inputMenu(habitList + ["all"], numbered=True)
    if habitToBeUpdated == "all":
        for habitElement in range(len(habitList)):
            yesOrNo = pyip.inputYesNo(
                f"Did you finish "
                f"{habitList[habitElement]}\n", yesVal="y", noVal="n"
            )
            sheetHabit.cell(
                row=findHabitrow(habitList[habitElement]), column=col
            ).value = yesOrNo
    else:
        yesOrNo = pyip.inputYesNo(
            f"Did you finish "
            f"{habitToBeUpdated}\n", yesVal="y", noVal="n"
        )
        sheetHabit.cell(
        row=findHabitrow(habitToBeUpdated), column=col
        ).value = yesOrNo
    print("Creating backup...")
    backupFolder = "Workbook Backups"
    os.makedirs(backupFolder, exist_ok=True)
    shutil.move(
        os.path.abspath(workbookChoice),
        os.path.abspath(backupFolder) + "//" + workbookChoice,
    )
    print("Writing...")
    update.wbVersion = (
#                                           Create the name of the file according to the current date
        "Habit Tracker" + "-" + \
        str(today.month) + "_" + str(today.day) + ".xlsx"
    )
    wb.save(update.wbVersion)
#                                           Use send2trash module instead of shutil.rm
#                                           to prevent from permanentaly removing the file
    from send2trash import send2trash as s2t

    workbookList = os.listdir("Workbook Backups")
    workbookList.sort()
    if not len(workbookList) > 3:
        return
    print("Deleting files older than three days")
    for i in range(3, len(workbookList)):
        s2t(os.path.join(
        os.path.abspath("Workbook Backups"), workbookList[i]))
